import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
import datetime as dt
from GLL_List_load import *


today = dt.date.today()
q = ((today.month - 1) // 3) + 1
y = today.year

prev_q = 4 if q == 1 else q - 1
prev_y = y - 1 if q == 1 else y

this_qtr = f"Q{q} {y}"
prev_qtr = f"Q{prev_q} {prev_y}"

scenario = "Baringa RC"

#store workbook
comp_file = "Q1 2026 Result Comparison.xlsx"
wb_comparison = openpyxl.load_workbook(comp_file, data_only=True)

GLL_file = "Q1 26 - Grid Model Forecast - Multi Scenario Internal_final.xlsx"
wb_GLL = openpyxl.load_workbook(GLL_file, data_only=True)

ws_name = f"GLL - {scenario}"
ws_GLL = wb_GLL[ws_name]

#convert MLF and curt table in GLL sheet
df_GLL_MLF = pd.read_excel(
    GLL_file,
    sheet_name=ws_name,
    header=4,     #Excel row 5
    skiprows=[5]  #data starts at row 7
)

df_GLL_curt = pd.read_excel(
    GLL_file,
    sheet_name=ws_name,
    header=31,     #Excel row 32
    skiprows=[32]  #data starts effectively at row 34
)


#find end of MLF table
col_name_MLF = df_GLL_MLF.iloc[:, 1]
mask_empty = col_name_MLF.isna() | (col_name_MLF.astype(str).str.strip() == "")


first_blank_idx = mask_empty.idxmax()
df_GLL_MLF = df_GLL_MLF.iloc[:first_blank_idx]


#modify columns
def modify_columns (df_GLL):
    df_GLL = df_GLL.rename(columns={"Project": "Name"})

    cols = [col for col in df_GLL.columns if str(col).startswith("20")]
    modified = df_GLL[["Name"] + cols]
    return modified

df_GLL_MLF = modify_columns(df_GLL_MLF)
df_GLL_curt = modify_columns(df_GLL_curt)


#create a new sheet in comparison workbook
def get_or_create_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    else:
        return wb.create_sheet(sheet_name)

sheet_name = f"{this_qtr} - {scenario}"

ws_this_qtr = get_or_create_sheet(wb_comparison, sheet_name)

#assign values to cells
ws_this_qtr['A1'] = "MLF"

header_row_MLF = 2

#column headers
def write_headers (header_row):
    ws_this_qtr.cell(row=header_row, column=1, value="State").font = Font(bold=True)
    ws_this_qtr.cell(row=header_row, column=1).alignment = Alignment(horizontal="center")

    ws_this_qtr.cell(row=header_row, column=2, value="Name").font = Font(bold=True)
    ws_this_qtr.cell(row=header_row, column=2).alignment = Alignment(horizontal="center")

write_headers(header_row_MLF)


#write FY columns
def write_FY_cols (header_row):
    start_year = base_year + 6
    end_year = base_year + 45

    FY_col_start = 3
    for year in range(start_year, end_year + 1):
        cell = ws_this_qtr.cell(row=header_row, column=FY_col_start, value=year)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        FY_col_start += 1

write_FY_cols(header_row_MLF)

#write name and states
start_row_MLF = header_row_MLF+1

def write_names_and_states(start_row,ws_name):
    for i, (name,state ) in enumerate(zip(name_list, state_list), start=start_row):
        ws_name.cell(row=i, column=2, value=name)  # Column B
        ws_name.cell(row=i, column=1, value=state)  # Column A

write_names_and_states(start_row_MLF,ws_this_qtr)


wb_comparison.save(comp_file)

#convert Excel data to dataframe
df_thisQ_MLF = pd.read_excel(
    comp_file,
    sheet_name=sheet_name,
    header=1,     # Excel row 2
)

#merge dataframes
def merge_df(df_GLL,df_thisQ):
    #drop columns that don't intersect
    df_GLL= df_GLL[df_GLL.columns.intersection(df_thisQ.columns)]

    #index match
    source_i = df_GLL.set_index("Name")
    target_i = df_thisQ.set_index("Name")
    target_i.update(source_i)
    df_thisQ = target_i.reset_index()

    cols = df_thisQ.columns.tolist()
    i, j = cols.index("Name"), cols.index("State")
    cols[i], cols[j] = cols[j], cols[i]
    merged = df_thisQ[cols]

    return merged

df_thisQ_MLF = merge_df(df_GLL_MLF,df_thisQ_MLF)

def write_df_excel(header_row,df_thisQ):
    # Write df starting at row 2 (Excel row 2)
    start_row = header_row # row 1 is preserved
    start_col = 1  # column A
    # Write column headers first
    for j, col_name in enumerate(df_thisQ.columns, start=start_col):
       ws_this_qtr.cell(row=start_row, column=j, value=col_name)
    # Write data
    for i, row in enumerate(df_thisQ.itertuples(index=False), start=start_row + 1):
       for j, value in enumerate(row, start=start_col):
           ws_this_qtr.cell(row=i, column=j, value=value)


write_df_excel(header_row_MLF,df_thisQ_MLF)

#save workbook
wb_comparison.save(comp_file)

df_thisQ_MLF = pd.read_excel(
    comp_file,
    sheet_name=sheet_name,
    header=1,     # Excel row 2
)

col_name = df_thisQ_MLF["Name"]

#find first blank cell in Name
mask_empty = col_name.isna() | (col_name.astype(str).str.strip() == "")
if mask_empty.any():
   first_blank_idx = mask_empty.idxmax()
else:
   first_blank_idx = len(df_thisQ_MLF)  #place at end if no blanks

#excel row = pandas index + header row + 1
header_row_excel = 2  #header in Excel row 2
curt_row_idx = first_blank_idx + header_row_excel + 2 #find first blank row index in pandas

ws_this_qtr[f'A{curt_row_idx}'] = "Curtailment"

header_row_curt = curt_row_idx + 1
write_headers(header_row_curt)
write_FY_cols(header_row_curt)

start_row_curt = header_row_curt+1
write_names_and_states(start_row_curt,ws_this_qtr)

wb_comparison.save(comp_file)

print(header_row_curt)

#convert Excel data to dataframe
df_thisQ_curt = pd.read_excel(
    comp_file,
    sheet_name=sheet_name,
    header=curt_row_idx,
)

print(df_thisQ_curt)


df_thisQ_curt = merge_df(df_GLL_curt,df_thisQ_curt)
write_df_excel(header_row_curt,df_thisQ_curt)

#Save workbook
wb_comparison.save(comp_file)

print(df_GLL_curt)

