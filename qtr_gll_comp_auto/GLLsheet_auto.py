import pandas as pd
import openpyxl
from GLL_List_load import *

#store workbook
wb_summaryOut = "Q1 2026 - Summary - Baringa RC.xlsx"

#open worksheet
ws_summaryOut_mlf = pd.read_excel(wb_summaryOut,sheet_name="mlf_summary")
ws_summaryOut_curt = pd.read_excel(wb_summaryOut,sheet_name="curt_summary")

#load the GLL sheet template
file_path = "Q1 26 - Grid Model Forecast - Multi Scenario Internal_final.xlsx"
wb_gll = openpyxl.load_workbook(file_path, data_only=True)

#select the GLL sheet
ws_gll = wb_gll["GLL - Baringa RC"]
name_col = 2   # Name in column B

#GLL Sheet - MLF section
start_row_mlf = 7
gll_header_mlf = 5

#GLL Sheet - Curt section
start_row_curt = start_row_mlf + 27
gll_header_curt = gll_header_mlf + 27


#filter summary output worksheet
def filter_summary_output(ws_summaryOut):
    ws_summaryOut["UniqueID"] = (
        ws_summaryOut["BusNum"].astype(str) + "_" +
        ws_summaryOut["GenID"].astype(str)
    )
    return ws_summaryOut[ws_summaryOut["UniqueID"].isin(genList["UniqueID"])].copy()

summaryMLF = filter_summary_output(ws_summaryOut_mlf)
summaryCurt = filter_summary_output(ws_summaryOut_curt)

#drop unwanted columns in summaryCurt
genList_cols = ["UniqueID", "BusNum", "GenID", "Name"]
curt_cols = [col for col in summaryCurt.columns if col.startswith("%_curt") and col.replace('%_curt','').isdigit()]

summaryCurt = summaryCurt[genList_cols + curt_cols]

#Column identifiers for in summaryMLF and summaryCurt
mlf_cols_start = "MLF"
curt_cols_start = "%_curt"

#merge genList with summary output sheet
def merge_genList_summary(summaryOut, cols_start):

    summary_cols = [col for col in summaryOut.columns if col not in genList_cols]
    df_merged = genList.merge(summaryOut[["UniqueID"] + summary_cols], on="UniqueID", how="left")
    #rename columns in summary output sheet to the respectives FY
    rename_dict = {
        col: f"{base_year + int(col.replace(cols_start, ''))}"
        for col in df_merged.columns
        if col.startswith(cols_start)
    }
    df_merged.rename(columns=rename_dict, inplace=True)
    return df_merged

genList_MLF = merge_genList_summary(summaryMLF, mlf_cols_start)
genList_Curt = merge_genList_summary(summaryCurt, curt_cols_start)


#write names and tech in MLF and Curt table
def write_names_and_tech(start_row):
    for i, (name, tech) in enumerate(zip(name_list, tech_list), start=start_row):
        ws_gll.cell(row=i, column=2, value=name)  # Column B
        ws_gll.cell(row=i, column=3, value=tech)  # Column C


write_names_and_tech(start_row_mlf)
write_names_and_tech(start_row_curt)


#calculate australian financial year
def get_aus_fy (cod):
    if cod.month >=7:
        return cod.year + 1
    else:
        return cod.year


def fill_forward(summaryOutput, start_row, gll_header):
    year_cols = []
    for col in range(1, ws_gll.max_column + 1):
        header = ws_gll.cell(row=gll_header, column=col).value
        if header and str(header).isdigit() and len(str(header)) == 4:
            year_cols.append((int(header), col))

    #sort the year columns
    year_cols.sort(key=lambda x: x[0])

    def get_year_value(row, y):
        return row.get(y, row.get(str(y), None))

    for i, gen_row in summaryOutput.iterrows():
        cod = gen_row["COD"]
        if pd.isna(cod):
            continue

        fy_year = get_aus_fy(cod)
        fy_start = fy_year
        fy_end = fy_year + 34

        fy_missing = pd.isna(get_year_value(gen_row, fy_year))

        #find the first value after COD+1
        first_future_year = None
        first_future_value = None
        for y, _ in year_cols:
            if y < fy_start or y > fy_end:
                continue
            v = get_year_value(gen_row, y)
            if not pd.isna(v):
                first_future_year = y
                first_future_value = v
                break

        last_value = None
        excel_row = start_row + i

        for year, col_idx in year_cols:
            cell = ws_gll.cell(row=excel_row, column=col_idx)

            if fy_start <= year <= fy_end:
                value = get_year_value(gen_row, year)

                if not pd.isna(value):

                    cell.value = value
                    last_value = value
                else:
                    #backfill values from first value to fy_start
                    if (
                        fy_missing
                        and first_future_year is not None
                        and fy_start <= year < first_future_year
                    ):
                        cell.value = first_future_value
                        last_value = first_future_value
                    else:
                        #forward fill values if not in the summary output
                        cell.value = last_value if last_value is not None else None
            else:
                cell.value = "-" #put "-" for years outside the range

fill_forward(genList_MLF, start_row_mlf, gll_header_mlf)
fill_forward(genList_Curt, start_row_curt, gll_header_curt)


#save GLL sheet
wb_gll.save(file_path)


# print(genList_MLF)
# print(genList_Curt)
