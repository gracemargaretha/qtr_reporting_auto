import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import DataBarRule, FormatObject, CellIsRule, Rule
from openpyxl.utils import get_column_letter
import datetime as dt
from GLL_List_load import *

today = dt.date.today()
q = ((today.month - 1) // 3) + 1
y = today.year

prev_q = 4 if q == 1 else q - 1
prev_y = y - 1 if q == 1 else y

this_qtr = f"Q{q} {y}"
prev_qtr = f"Q{prev_q} {prev_y}"


scenario = "Baringa RC"

#store workbook and worksheet
comp_file = "Q1 2026 Result Comparison.xlsx"
wb_comparison = openpyxl.load_workbook(comp_file, data_only=True)

ws_this_qtr_name = f"{this_qtr} - {scenario}"
ws_prev_qtr_name = f"{prev_qtr} - {scenario}"

ws_this_qtr = wb_comparison[ws_this_qtr_name]
ws_prev_qtr = wb_comparison[ws_prev_qtr_name]

#define headers and start row
thisQ_header_MLF = prevQ_header_MLF = 2
thisQ_start_row_MLF = prevQ_start_row_MLF = thisQ_header_MLF + 1

#convert MLF and curt table in this and prev Q sheet
df_thisQ = pd.read_excel(
    comp_file,
    sheet_name=ws_this_qtr_name,
    header=thisQ_header_MLF-1
)

df_prevQ= pd.read_excel(
    comp_file,
    sheet_name=ws_prev_qtr_name,
    header=prevQ_header_MLF-1
)

#split MLF and Curt table
def split_df(df):

    #find row that contains 'Curtailment'
    curtailment_idx = df.apply(
        lambda row: row.astype(str).str.strip().str.lower().eq('curtailment').any(),
        axis=1
    ).idxmax()

    #split into MLF and Curtailment,keeping headers
    mlf_df = pd.concat(
        [pd.DataFrame([df.columns], columns=df.columns), df.iloc[:curtailment_idx]],
        ignore_index=True
    )

    curtailment_df = pd.concat(
        [pd.DataFrame([df.columns], columns=df.columns), df.iloc[curtailment_idx + 1:]],
        ignore_index=True
    )

    #remove repeated headers in Curtailment
    curtailment_df = curtailment_df[curtailment_df['State'] != 'State'].reset_index(drop=True)
    mlf_df = mlf_df[mlf_df['State'] != 'State'].reset_index(drop=True)

    return mlf_df, curtailment_df


df_thisQ_MLF, df_thisQ_curt = split_df(df_thisQ)
df_prevQ_MLF, df_prevQ_curt = split_df(df_prevQ)

#find common columns
df_thisQ_MLF = df_thisQ_MLF.dropna(how='all')
df_prevQ_MLF = df_prevQ_MLF.dropna(how='all')

common_MLF = sorted(set(df_prevQ_MLF['Name']).intersection(df_thisQ_MLF['Name']))
common_curt = sorted(set(df_prevQ_curt['Name']).intersection(df_thisQ_curt['Name']))


df_thisQ_MLF = (

    df_thisQ_MLF[df_thisQ_MLF['Name'].isin(common_MLF)]

    .sort_values(['State', 'Name'])

    .reset_index(drop=True)

)

df_prevQ_MLF = (

    df_prevQ_MLF[df_prevQ_MLF['Name'].isin(common_MLF)]

    .sort_values(['State', 'Name'])

    .reset_index(drop=True)

)

df_thisQ_curt = (

    df_thisQ_curt[df_thisQ_curt['Name'].isin(common_curt)]

    .sort_values(['State', 'Name'])

    .reset_index(drop=True)

)

df_prevQ_curt = (

    df_prevQ_curt[df_prevQ_curt['Name'].isin(common_curt)]

    .sort_values(['State', 'Name'])

    .reset_index(drop=True)

)


#calculate differences
def calculate_diff(df_current, df_prev):
    import pandas as pd

    # Identify year columns
    fy_cols = [c for c in df_current.columns if c not in ['Name', 'State']]
    # Convert to numeric and subtract
    diff_values = (
            df_current[fy_cols].apply(pd.to_numeric, errors='coerce')
            - df_prev[fy_cols].apply(pd.to_numeric, errors='coerce')
    )
    # Replace NaN with "-"
    diff_values = diff_values.fillna('-')
    # Build final df
    df_diff = df_current[['Name', 'State']].copy()
    df_diff[fy_cols] = diff_values
    return df_diff


df_diff_MLF = calculate_diff(df_thisQ_MLF, df_prevQ_MLF)
df_diff_curt = calculate_diff(df_thisQ_curt, df_prevQ_curt)

def calc_avg (df_diff):
    id_cols = ['Name','State']
    fy_cols = [c for c in df_diff.columns if c not in id_cols]

    df_diff_numeric = df_diff[fy_cols].apply(pd.to_numeric, errors='coerce')
    #calculate row-wise average
    avg_series = df_diff_numeric.mean(axis=1)
    #insert average column
    df_diff.insert(2, 'Average', avg_series)

    return df_diff

df_diff_MLF = calc_avg(df_diff_MLF)
df_diff_curt = calc_avg(df_diff_curt)


print(common_MLF)
print(common_curt)

#create a new sheet in comparison workbook
sheet_name = f"{this_qtr} vs {prev_qtr} {scenario}"

if sheet_name in wb_comparison.sheetnames:
    wb_comparison[sheet_name]
else:
    wb_comparison.create_sheet(sheet_name)

wb_comparison.save(comp_file)

#write a title in A1
ws_comparison =  wb_comparison[sheet_name]

ws_comparison['A1'] = ("MLF")


#write headers in row 2 (A2)
for col_idx, col_name in enumerate(df_diff_MLF.columns, 1):
    ws_comparison.cell(row=2, column=col_idx, value=col_name)

#write df starting from A3
for r_idx, row in enumerate(df_diff_MLF.values, start=3):
    for c_idx, value in enumerate(row, start=1):
        cell = ws_comparison.cell(row=r_idx, column=c_idx, value=value)
        # Format as percentage with 2 decimals
        cell.number_format = "0.00%"


wb_comparison.save(comp_file)

col_name = df_diff_MLF["Name"]

#find first blank cell in Name
mask_empty = col_name.isna() | (col_name.astype(str).str.strip() == "")
if mask_empty.any():
   first_blank_idx = mask_empty.idxmax()
else:
   first_blank_idx = len(df_diff_MLF)  #place at end if no blanks

#Excel row = pandas index + header row + 1
header_row_excel = 2  # header is in Excel row 2
curt_row_idx = first_blank_idx + header_row_excel + 2 # find first blank row index in pandas

ws_this_qtr[f'A{curt_row_idx}'] = "Curtailment"

#write the DataFrame headers
for col_idx, col_name in enumerate(df_diff_MLF.columns, 1):
    ws_comparison.cell(row=curt_row_idx+1, column=col_idx, value=col_name)

#write df starting from A3 (so headers remain on row 2)
for r_idx, row in enumerate(df_diff_curt.values, start=curt_row_idx+2):
    for c_idx, value in enumerate(row, start=1):
        cell = ws_comparison.cell(row=r_idx, column=c_idx, value=value)
        # Format as percentage with 2 decimals
        cell.number_format = "0.00%"


wb_comparison.save(comp_file)